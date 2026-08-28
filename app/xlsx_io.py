"""Excel template/export + upload parsing for the Specialty Shows calendar.

Two upload formats are understood:

1. Our round-trip template (built by build_template / build_export below) —
   one clean column per field, plus an EVENT ID column for in-place updates.

2. The master calendar workbook the show side already maintains
   ("ESS Specialty & Supported Entry Calendar": MONTH/DATES, SHOW TYPE,
   HOST CLUB, LOCATION, CLOSING DATE, SUPERINTENDENT, REGULAR JUDGE,
   SWEEPSTAKES/FUTURITY JUDGE). That file is messy in stable, predictable
   ways, and the parser handles each one:
     - month banner rows (JANUARY … DECEMBER) mark the section
     - a tail section previews early next year, marked by first-of-month
       dates or "March 2027"-style rows
     - date cells were rolled forward from last year, so the TYPED YEAR is
       unreliable — month/day are trusted, the year comes from the section
     - blank cells inherit from the row above (location and closing date are
       written once per cluster; a dateless row is a second show that day)
     - string dates like "4-Sept" / "1-Nov." / "14- Oct, 15- Oct, 16- Oct,
       17- Oct" (a range) appear alongside real date cells
   Every normalization is recorded as a per-row note so a human can review
   what the parser decided.
"""
import io
import re
from datetime import date, datetime, timedelta

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

from .db import SHOW_TYPES

BRAND = "681E12"

MONTH_NAMES = ["JANUARY", "FEBRUARY", "MARCH", "APRIL", "MAY", "JUNE", "JULY",
               "AUGUST", "SEPTEMBER", "OCTOBER", "NOVEMBER", "DECEMBER"]
MONTH_PREFIX = {m[:3].lower(): i + 1 for i, m in enumerate(MONTH_NAMES)}

HEADERS = [
    ("FIRST DAY (e.g. 2026-09-12)", 20),
    ("LAST DAY (blank = one-day)", 20),
    ("SHOW TYPE", 20),
    ("HOST CLUB", 36),
    ("CITY", 16),
    ("STATE", 8),
    ("VENUE / GROUNDS", 30),
    ("CLOSING DATE", 14),
    ("SUPERINTENDENT", 22),
    ("REGULAR JUDGE", 22),
    ("SWEEPS / FUTURITY JUDGE", 24),
]

EXAMPLE = ["2026-09-12", "", "Specialty", "EXAMPLE — ESS Club of Anytown (delete this row)",
           "Anytown", "OH", "County Fairgrounds", "2026-08-26", "Onofrio",
           "Jane Judge", "John Sweeps"]


def _norm(v):
    if v is None:
        return ""
    return " ".join(str(v).replace("\xa0", " ").split())


def _empty(v):
    return _norm(v) == ""


def norm_show_type(raw):
    """Map the sheet's spelling variants onto the canonical list.
    Returns (canonical_or_None, note_or_None)."""
    t = _norm(raw).lower().rstrip(".")
    if not t:
        return None, None
    t = t.replace("speciality", "specialty")
    if "group" in t and "supported" in t:
        return "Supported Entry", f"show type '{_norm(raw)}' recorded as Supported Entry (held with a group show)"
    for canon in SHOW_TYPES:
        if t == canon.lower():
            return canon, None
    if t == "supported":
        return "Supported Entry", None
    if "national" in t:
        return "National Specialty", None
    if "designated" in t:
        return "Designated Specialty", None
    if "concurrent" in t:
        return "Concurrent Specialty", None
    if "supported" in t:
        return "Supported Entry", None
    if "group" in t:
        return "Group Show", None
    if "specialty" in t:
        return "Specialty", None
    return "Specialty", f"unrecognized show type '{_norm(raw)}' recorded as Specialty"


STATE_ABBREV = {"IND": "IN", "ILL": "IL", "FLA": "FL", "MASS": "MA", "CONN": "CT",
                "TEX": "TX", "CALIF": "CA", "MICH": "MI", "WISC": "WI", "MINN": "MN",
                "PENN": "PA", "ORE": "OR", "WASH": "WA", "COLO": "CO", "ARIZ": "AZ"}


def split_location(raw):
    """'Venue name\nCity, ST' / 'City, ST' / long free text → (venue, city, state).
    The state is a trailing two-letter (or known longer) abbreviation; the city is
    only claimed when it is unambiguous, otherwise the text stays in venue."""
    if _empty(raw):
        return "", "", ""
    lines = [" ".join(l.split()) for l in str(raw).replace("\xa0", " ").splitlines() if l.strip()]
    # runs of 3+ spaces inside one visual line act as line breaks in the sheet
    if len(lines) == 1:
        lines = [p.strip() for p in re.split(r"\s{3,}", str(raw).replace("\xa0", " ").strip()) if p.strip()]
        lines = [" ".join(l.split()) for l in lines]
    last = lines[-1]
    state = ""
    m = re.search(r"[,\s]([A-Za-z]{2,5})\.?\s*$", " " + last)
    if m:
        cand = m.group(1).upper().rstrip(".")
        if len(cand) == 2 and cand.isalpha():
            state = cand
        elif cand in STATE_ABBREV:
            state = STATE_ABBREV[cand]
    if state:
        last = last[:m.start()].rstrip(" ,.")
    city = ""
    if last and "," not in last and len(last.split()) <= 4 and len(lines) > 1:
        city = last          # 'Venue…' lines above, plain 'City' line before the state
        lines = lines[:-1]
    elif last and len(last.split()) <= 4 and len(lines) == 1:
        city = last          # the whole cell was 'City, ST'
        lines = []
    elif last:
        lines[-1] = last     # long text: leave it in venue rather than guess a city
    venue = ", ".join(lines)
    return venue, city, state


def _parse_date_token(tok, year):
    """'4-Sept' / '1-Nov.' / '21-Feb' / 'Oct 14' → date in the given year, else None."""
    t = _norm(tok).rstrip(".").replace("–", "-")
    m = re.match(r"^(\d{1,2})\s*-?\s*([A-Za-z]{3,9})$", t) or \
        re.match(r"^([A-Za-z]{3,9})\.?\s*-?\s*(\d{1,2})$", t)
    if not m:
        return None
    a, b = m.group(1), m.group(2)
    day_s, mon_s = (a, b) if a.isdigit() else (b, a)
    mon = MONTH_PREFIX.get(mon_s[:3].lower())
    if not mon:
        return None
    try:
        return date(year, mon, int(day_s))
    except ValueError:
        return None


def parse_show_date(v, section_year, section_month):
    """The MONTH/DATES cell → (start, end, note) with the year taken from the
    section, not from what was typed. Returns (None, None, reason) on failure."""
    if isinstance(v, (datetime, date)):
        d = v.date() if isinstance(v, datetime) else v
        year = section_year
        if section_month == 12 and d.month == 1:
            year = section_year + 1  # January show listed at the year's end
        fixed = date(year, d.month, d.day)
        note = None
        if d.year != fixed.year:
            note = f"typed year {d.year} corrected to {fixed.year}"
        if section_month and d.month != section_month and not (section_month == 12 and d.month == 1):
            note = ((note + "; ") if note else "") + \
                f"date month ({MONTH_NAMES[d.month-1].title()}) differs from the sheet section ({MONTH_NAMES[section_month-1].title()}) — check"
        return fixed, fixed, note
    s = _norm(v)
    if not s:
        return None, None, None
    parts = [p for p in re.split(r"[,;/]", s) if p.strip()]
    dates = [_parse_date_token(p, section_year) for p in parts]
    dates = [d for d in dates if d]
    if not dates:
        return None, None, f"unreadable date '{s}'"
    start, end = min(dates), max(dates)
    note = f"date '{s}' read as {start.isoformat()}" + (f" – {end.isoformat()}" if end != start else "")
    return start, end, note


def parse_closing_date(v, start):
    """Closing-date cell → (iso_or_'', note). Typed years are unreliable; the
    month/day are re-anchored to the year that puts closing shortly before the show."""
    if _empty(v) or _norm(v) in ("?", "-", "N/A", "TBD"):
        return "", None
    if isinstance(v, (datetime, date)):
        d = v.date() if isinstance(v, datetime) else v
        mon, day = d.month, d.day
        typed_year = d.year
    else:
        parsed = _parse_date_token(_norm(v), start.year)
        if not parsed:
            return "", f"unreadable closing date '{_norm(v)}'"
        mon, day, typed_year = parsed.month, parsed.day, None
    best = None
    for y in (start.year, start.year - 1):
        try:
            cand = date(y, mon, day)
        except ValueError:
            continue
        gap = (start - cand).days
        if 0 <= gap <= 180 and (best is None or gap < (start - best).days):
            best = cand
    if best is None:
        try:
            best = date(start.year, mon, day)
        except ValueError:
            return "", f"impossible closing date '{_norm(v)}'"
        note = f"closing date {best.isoformat()} is not shortly before the show — check"
        return best.isoformat(), note
    note = None
    if typed_year is not None and typed_year != best.year:
        note = f"closing-date year {typed_year} corrected to {best.year}"
    return best.isoformat(), note


def _judge(v):
    s = _norm(v)
    return "" if s.upper() in ("N/A", "NA", "?", "-", "") else s


def _is_blankish(row):
    return all(_empty(c) for c in row)


def parse_master(ws, calendar_year, id_col=None):
    """The master-calendar format. Returns (rows, errors). Each row dict carries
    a '_notes' list describing what the parser normalized on that row."""
    rows, errors = [], []
    header_row = None
    for r_i, row in enumerate(ws.iter_rows(max_row=6, values_only=True), 1):
        texts = [_norm(c).upper() for c in row]
        if any("MONTH" in t and "DATE" in t for t in texts):
            header_row = r_i
            break
    if header_row is None:
        return [], ["Could not find the MONTH/DATES header row — is this the specialty calendar workbook?"]

    section_year, section_month = calendar_year, None
    prev = None
    for r_i, row in enumerate(ws.iter_rows(min_row=header_row + 1, values_only=True), header_row + 1):
        if _is_blankish(row):
            continue
        first = row[0]
        rest_empty = all(_empty(c) for c in row[1:])
        # ---- section markers ----
        s0 = _norm(first).upper().rstrip(".")
        if s0 in MONTH_NAMES and rest_empty:
            section_month = MONTH_NAMES.index(s0) + 1
            continue
        if isinstance(first, (datetime, date)) and rest_empty:
            d = first.date() if isinstance(first, datetime) else first
            if d.day == 1:  # '2027-01-01' style next-year marker
                section_year, section_month = d.year, d.month
                continue
        m = re.match(r"^([A-Za-z]{3,9})\.?\s+(\d{4})$", _norm(first))
        if m and rest_empty:
            mon = MONTH_PREFIX.get(m.group(1)[:3].lower())
            if mon:
                section_year, section_month = int(m.group(2)), mon
                continue
        # ---- data row ----
        raw_type, raw_club, raw_loc = row[1], row[2], row[3]
        raw_close, raw_super = row[4], row[5]
        raw_j1, raw_j2 = (row[6] if len(row) > 6 else None), (row[7] if len(row) > 7 else None)
        notes = []
        start, end, note = parse_show_date(first, section_year, section_month or 1)
        if note and start is None:
            errors.append(f"Row {r_i}: {note} — skipped")
            continue
        if note:
            notes.append(note)
        if start is None:
            if prev is None:
                continue  # stray fragment before any complete row
            start = date.fromisoformat(prev["start_date"])
            end = date.fromisoformat(prev["end_date"])
            notes.append("no date — read as another show on the same day as the row above")
        show_type, tnote = norm_show_type(raw_type)
        if tnote:
            notes.append(tnote)
        club = _norm(raw_club)
        if show_type is None:
            if not club and prev:
                show_type = prev["show_type"]
            else:
                show_type, tnote = (prev["show_type"], "show type inherited from the row above") if prev else ("Specialty", "no show type — recorded as Specialty")
                notes.append(tnote)
        if not club:
            if prev is None:
                errors.append(f"Row {r_i}: no host club and nothing above to inherit — skipped")
                continue
            club = prev["club"]
            notes.append("host club inherited from the row above")
        if not _empty(raw_loc):
            venue, city, state = split_location(raw_loc)
        elif prev is not None:
            venue, city, state = prev["venue"], prev["city"], prev["state"]
            if venue or city or state:
                notes.append("location inherited from the row above")
        else:
            venue = city = state = ""
        if not _empty(raw_close):
            closing, cnote = parse_closing_date(raw_close, start)
            if cnote:
                notes.append(cnote)
        else:
            closing = prev["closing_date"] if prev else ""
        superintendent = _norm(raw_super).lstrip("?").strip() if not _empty(raw_super) else (prev["superintendent"] if prev else "")
        if _norm(raw_super) == "?":
            superintendent = ""
        row_id = None
        if id_col is not None and id_col < len(row) and not _empty(row[id_col]):
            try:
                row_id = int(float(row[id_col]))
            except (TypeError, ValueError):
                errors.append(f"Row {r_i} ({club}): EVENT ID '{_norm(row[id_col])}' isn't a number — treated as a new event")
        rec = {
            "id": row_id,
            "title": club,
            "club": club,
            "show_type": show_type,
            "start_date": start.isoformat(),
            "end_date": end.isoformat(),
            "city": city, "state": state, "venue": venue,
            "closing_date": closing,
            "superintendent": superintendent,
            "judge_regular": _judge(raw_j1),
            "judge_sweeps": _judge(raw_j2),
            "status": "scheduled",
            "source": "xlsx-import",
            "_row": r_i,
            "_notes": notes,
        }
        rows.append(rec)
        prev = rec
    return rows, errors


def parse_template(ws, id_col_hint=None):
    """Our own round-trip format: explicit columns, header-driven."""
    header_row, cols = None, {}
    for r_i, row in enumerate(ws.iter_rows(max_row=10, values_only=True), 1):
        texts = [_norm(c).upper() for c in row]
        if any("CLUB" in t for t in texts):
            header_row = r_i
            for c_i, t in enumerate(texts):
                if not t:
                    continue
                if "EVENT ID" in t or t == "ID":
                    cols["id"] = c_i
                elif t.startswith("FIRST DAY"):
                    cols["start"] = c_i
                elif t.startswith("LAST DAY"):
                    cols["end"] = c_i
                elif "SHOW TYPE" in t or t == "TYPE":
                    cols["type"] = c_i
                elif "CLUB" in t:
                    cols["club"] = c_i
                elif "CITY" in t:
                    cols["city"] = c_i
                elif "STATE" in t:
                    cols["state"] = c_i
                elif "VENUE" in t or "GROUNDS" in t:
                    cols["venue"] = c_i
                elif "CLOSING" in t:
                    cols["closing"] = c_i
                elif "SUPER" in t:
                    cols["super"] = c_i
                elif "SWEEP" in t or "FUTURITY" in t:
                    cols["sweeps"] = c_i
                elif "JUDGE" in t:
                    cols["judge"] = c_i
            break
    if header_row is None or "club" not in cols or "start" not in cols:
        return [], ["Could not find a header row with HOST CLUB and FIRST DAY columns."]

    def _plain_date(v):
        if isinstance(v, (datetime, date)):
            return (v.date() if isinstance(v, datetime) else v).isoformat()
        s = _norm(v)
        if not s or s == "-":
            return None
        for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y", "%d %b %Y", "%b %d %Y", "%b %d, %Y"):
            try:
                return datetime.strptime(s, fmt).date().isoformat()
            except ValueError:
                pass
        return "ERR"

    rows, errors = [], []
    for r_i, row in enumerate(ws.iter_rows(min_row=header_row + 1, values_only=True), header_row + 1):
        get = lambda key: row[cols[key]] if key in cols and cols[key] < len(row) else None
        club = _norm(get("club"))
        if not club or club == "-" or "EXAMPLE" in club.upper():
            continue
        start = _plain_date(get("start"))
        end = _plain_date(get("end"))
        if start in (None, "ERR"):
            errors.append(f"Row {r_i} ({club}): missing or unreadable first-day date — skipped")
            continue
        if end in (None, "ERR"):
            end = start
        if end < start:
            start, end = end, start
        closing = _plain_date(get("closing"))
        if closing in (None, "ERR"):
            closing = ""
        show_type, tnote = norm_show_type(get("type"))
        row_id = None
        if "id" in cols and not _empty(get("id")):
            try:
                row_id = int(float(get("id")))
            except (TypeError, ValueError):
                errors.append(f"Row {r_i} ({club}): EVENT ID '{_norm(get('id'))}' isn't a number — treated as a new event")
        rows.append({
            "id": row_id,
            "title": club,
            "club": club,
            "show_type": show_type or "Specialty",
            "start_date": start,
            "end_date": end,
            "city": _norm(get("city")),
            "state": _norm(get("state")).upper()[:2],
            "venue": _norm(get("venue")),
            "closing_date": closing,
            "superintendent": _norm(get("super")),
            "judge_regular": _judge(get("judge")),
            "judge_sweeps": _judge(get("sweeps")),
            "status": "scheduled",
            "source": "xlsx-import",
            "_row": r_i,
            "_notes": ([tnote] if tnote else []),
        })
    return rows, errors


def parse_upload(data: bytes, calendar_year: int):
    """Returns (rows, errors). Detects which of the two formats the file is in."""
    wb = load_workbook(io.BytesIO(data), data_only=True)
    ws = wb.active
    id_col = None
    is_master = False
    for r_i, row in enumerate(ws.iter_rows(max_row=6, values_only=True), 1):
        texts = [_norm(c).upper() for c in row]
        if any("MONTH" in t and "DATE" in t for t in texts):
            is_master = True
            for c_i, t in enumerate(texts):
                if "EVENT ID" in t:
                    id_col = c_i
            break
        if any(t.startswith("FIRST DAY") for t in texts):
            break
    if is_master:
        return parse_master(ws, calendar_year, id_col=id_col)
    return parse_template(ws)


# ---------- template + export ----------

def build_template() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Specialty Shows"
    banner = ws.cell(row=1, column=1, value="ESS Specialty & Supported Entry Calendar — upload template")
    banner.font = Font(name="Arial", bold=True, size=13, color="FFFFFF")
    banner.fill = PatternFill("solid", fgColor=BRAND)
    banner.alignment = Alignment(horizontal="center")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(HEADERS))
    ws.append([h for h, _ in HEADERS])
    for i, (_, width) in enumerate(HEADERS, 1):
        ws.column_dimensions[ws.cell(row=2, column=i).column_letter].width = width
        c = ws.cell(row=2, column=i)
        c.font = Font(name="Arial", bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=BRAND)
        c.alignment = Alignment(horizontal="center", wrap_text=True)
    ws.append(EXAMPLE)
    for cell in ws[3]:
        cell.font = Font(name="Arial", italic=True, color="888888")
    ws.append([])
    note = ws.cell(row=5, column=1,
                   value="Fill one row per show (a back-to-back weekend is two rows). Show types: "
                         + " / ".join(SHOW_TYPES) + ". Days of the week are added automatically "
                         "from the dates. Delete the gray example row before uploading. "
                         "Dates: 2026-09-12 style, or use Excel date cells. You can also upload the "
                         "master 'ESS Specialty & Supported Entry Calendar' workbook itself — "
                         "the site understands that format too.")
    note.font = Font(name="Arial", italic=True, size=9)
    ws.merge_cells(start_row=5, start_column=1, end_row=5, end_column=len(HEADERS))
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_export(events, year: int) -> bytes:
    """The calendar's current shows in the template layout, plus an EVENT ID
    column so an edited re-upload can update events in place."""
    wb = Workbook()
    ws = wb.active
    ws.title = f"Specialty Shows {year}"[:31]
    banner = ws.cell(row=1, column=1, value=f"ESS Specialty & Supported Entry Calendar — {year}")
    banner.font = Font(name="Arial", bold=True, size=13, color="FFFFFF")
    banner.fill = PatternFill("solid", fgColor=BRAND)
    banner.alignment = Alignment(horizontal="center")
    all_headers = [h for h, _ in HEADERS] + ["STATUS (info only)", "EVENT ID (do not change)"]
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(all_headers))
    ws.append(all_headers)
    for i, width in enumerate([w for _, w in HEADERS] + [16, 24], 1):
        ws.column_dimensions[ws.cell(row=2, column=i).column_letter].width = width
        c = ws.cell(row=2, column=i)
        c.font = Font(name="Arial", bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=BRAND)
        c.alignment = Alignment(horizontal="center", wrap_text=True)
    for ev in events:
        ws.append([
            ev["start_date"][:10],
            ev["end_date"][:10] if ev["end_date"][:10] != ev["start_date"][:10] else "",
            ev["show_type"],
            ev["club"] or ev["title"],
            ev["city"], ev["state"], ev["venue"],
            ev["closing_date"][:10] if ev["closing_date"] else "",
            ev["superintendent"],
            ev["judge_regular"], ev["judge_sweeps"],
            ev["status"] + (" +hidden" if ev.get("hidden") else ""),
            ev["id"],
        ])
    note = ws.cell(row=ws.max_row + 2, column=1,
                   value="Edit any column and re-upload to UPDATE these shows in place "
                         "(you'll see every change before it's saved). Leave EVENT ID untouched. "
                         "Rows you add without an ID become new shows. To make NEXT year's calendar "
                         "from this one, change the dates and tick 'treat every row as new' when "
                         "uploading. The STATUS column is informational and ignored on upload.")
    note.font = Font(name="Arial", italic=True, size=9)
    ws.merge_cells(start_row=ws.max_row, start_column=1, end_row=ws.max_row, end_column=len(all_headers))
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
